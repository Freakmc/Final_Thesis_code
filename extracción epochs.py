import openpyxl
import numpy as np
import csv
import pandas as pd
import math
from scipy.signal import butter, filtfilt
import matplotlib.pyplot as plt

def visualizar(path_archivo_a_visualizar):

    datos = []
    tiempo = []
    canal_1 = []
    canal_2 = []
    canal_3 = []
    canal_4 = []

    with open(path_archivo_a_visualizar, "r") as archivo:
        for linea in archivo:
            datos.append(linea.strip().split(","))
    with open(path_archivo_a_visualizar, "r") as archivo:
        lector = csv.reader(archivo)
        next(lector)  # Saltar la primera fila si contiene encabezados
        for fila in lector:
            tiempo.append(float(fila[0]))
            canal_1.append(float(fila[1]))
            canal_2.append(float(fila[2]))
            canal_3.append(float(fila[3]))
            canal_4.append(float(fila[4]))

    #Plotear los canales
    marcas_tiempo= np.linspace(0,len(tiempo)/256,len(tiempo))
    plt.figure(figsize=(10, 8))
    plt.subplot(4, 1, 1)
    plt.plot(marcas_tiempo, canal_1, color='blue')
    plt.title("TP9")
    plt.xlabel("Tiempo")
    plt.ylabel("Amplitud")
    plt.subplot(4, 1, 2)
    plt.plot(marcas_tiempo, canal_2, color='green')
    plt.title("AF7")
    plt.xlabel("Tiempo")
    plt.ylabel("Amplitud")
    plt.subplot(4, 1, 3)
    plt.plot(marcas_tiempo, canal_3, color='red')
    plt.title("AF8")
    plt.xlabel("Tiempo")
    plt.ylabel("Amplitud")
    plt.subplot(4, 1, 4)
    plt.plot(marcas_tiempo, canal_4, color='purple')
    plt.title("TP10")
    plt.xlabel("Tiempo")
    plt.ylabel("Amplitud")
    plt.tight_layout()
    plt.show()

def posiciones(erp, ruta):
    # Cargar el archivo de origen
    wb = openpyxl.load_workbook(ruta)
    hoja = wb.active

    #Definir variable con las palabras y frases incongruentes
    palabras=["Caballo","Lujo","Salida","Teclado","Firma","Planeta","Aspirina","Microscopio",
        "Béisbol","Videojuego","Almohadón","Olas","Libro","Cascada","Tijera","Factura","Barco"]
    frases=[
        "Todas las noches bebo jamón",
        "Los peatones circulan por las acelgas",
        "El perro es el mejor amigo del calzoncillo",
        "Ayer me comí una llave inglesa",
        "El sastre me cosió las orejas",
        "Juan pela las gambas con el móvil"]
    caras=["face_r1.jpg","face_r2.jpg","face_r3.jpg","face_r4.jpg","face_r5.jpg",
        "face_r6.jpg","face_r7.jpg","face_r8.jpg","face_r9.jpg","face_r10.jpg"]
    
    # Definir las variables, fotos asignada y elegida
    if(hoja['GP2'].value.lower()=="pikachu"): foto_elegida = "s1_"+hoja['GP2'].value.lower()+".png"
    else: foto_elegida = "s1_"+hoja['GP2'].value.lower()+".jpg"
    foto_asignada = "s2_"+hoja['GQ2'].value.lower()+".jpg"
    #print(foto_elegida,foto_asignada) 

    # Creo arrays para sacar el orden de aparición de los elementos
    elementos = []
    posiciones = [] 

    # Dependiendo del protocolo cojo las celdas que correspondan
    if erp=="p300_1_1":
        for fila in hoja.iter_rows(min_row=6, max_row=41, min_col=1, max_col=1):
            for celda in fila:
                elementos.append(celda.value)
        #print(elementos)
        # Guardo las posiciones en las que se encuentra la palabra que busco
        posiciones = [i+1 for i, x in enumerate(elementos) if x == foto_elegida] 
        #print(posiciones)
    if erp=="p300_2_1":
        for fila in hoja.iter_rows(min_row=43, max_row=79, min_col=1, max_col=1):
            for celda in fila:
                elementos.append(celda.value)
        #print(elementos)
        posiciones = [i+1 for i, x in enumerate(elementos) if x == foto_asignada]
        #print(posiciones)
    if erp=="p300_1_2":
        for fila in hoja.iter_rows(min_row=81, max_row=114, min_col=1, max_col=1):
            for celda in fila:
                elementos.append(celda.value)
        #print(elementos)
        posiciones = [i+1 for i, x in enumerate(elementos) if x == foto_elegida]
        #print(posiciones)
    if erp=="p300_2_2":
        for fila in hoja.iter_rows(min_row=116, max_row=150, min_col=1, max_col=1):
            for celda in fila:
                elementos.append(celda.value)
        #print(elementos)
        posiciones = [i+1 for i, x in enumerate(elementos) if x == foto_asignada]
        #print(posiciones)
    if erp=="n400_1":
        for fila in hoja.iter_rows(min_row=153, max_row=181, min_col=3, max_col=3):
            for celda in fila:
                elementos.append(celda.value)
        #print(elementos)
        for palabra in palabras:
            posiciones.extend([i+1 for i, x in enumerate(elementos) if x == palabra])
        posiciones = sorted(set(posiciones))
        #print(posiciones)
    if erp=="n400_2":
        for fila in hoja.iter_rows(min_row=183, max_row=200, min_col=4, max_col=4):
            for celda in fila:
                elementos.append(celda.value)
        #print(elementos)
        for frase in frases:
            posiciones.extend([i+1 for i, x in enumerate(elementos) if x == frase])
        posiciones = sorted(set(posiciones))
        #print(posiciones)
    if erp=="p300_1_3":
        for fila in hoja.iter_rows(min_row=204, max_row=240, min_col=1, max_col=1):
            for celda in fila:
                elementos.append(celda.value)
        #print(elementos)
        posiciones = [i+1 for i, x in enumerate(elementos) if x == "s1_"+hoja['GP2'].value.lower()+".jpg"]
        #print(posiciones)
    if erp=="p300_2_3":
        for fila in hoja.iter_rows(min_row=242, max_row=273, min_col=1, max_col=1):
            for celda in fila:
                elementos.append(celda.value)
        #print(elementos)
        posiciones = [i+1 for i, x in enumerate(elementos) if x == foto_asignada]
        #print(posiciones)
    if erp=="p300_1_4":
        for fila in hoja.iter_rows(min_row=275, max_row=306, min_col=1, max_col=1):
            for celda in fila:
                elementos.append(celda.value)
        #print(elementos)
        posiciones = [i+1 for i, x in enumerate(elementos) if x == "s1_"+hoja['GP2'].value.lower()+".jpg"]
        #print(posiciones)
    if erp=="p300_2_4":
        for fila in hoja.iter_rows(min_row=308, max_row=344, min_col=1, max_col=1):
            for celda in fila:
                elementos.append(celda.value)
        #print(elementos)
        posiciones = [i+1 for i, x in enumerate(elementos) if x == foto_asignada]
        #print(posiciones)
    if erp=="n400_3":
        for fila in hoja.iter_rows(min_row=347, max_row=375, min_col=3, max_col=3):
            for celda in fila:
                elementos.append(celda.value)
        #print(elementos)
        for palabra in palabras:
            posiciones.extend([i+1 for i, x in enumerate(elementos) if x == palabra])
        posiciones = sorted(set(posiciones))
        #print(posiciones)
    if erp=="n400_4":
        posiciones = [3,6,7]
        #print(posiciones)
    if erp=="n400_5":
        for fila in hoja.iter_rows(min_row=379, max_row=408, min_col=5, max_col=5):
            for celda in fila:
                elementos.append(celda.value)
        #print(elementos)
        for cara in caras:
            posiciones.extend([i+1 for i, x in enumerate(elementos) if x == cara])
        posiciones = sorted(set(posiciones))
        #print(posiciones)
    print("se han detectado "+ str(len(posiciones)) + " potenciales")
    return posiciones

def extraer_epoca(ruta_extracción,posiciones,erps):

    df = pd.read_csv(ruta_extracción) # Cargar el archivo CSV
    num_muestras = len(df) # Calcular el número de muestras
    fs = 256  # Calcular el tiempo de grabado (Hz)
    Tiempo_grabado=num_muestras/fs

    Tiempos_reales=[36,37,34,35,37.7,72,37,32,32,37,37.7,20,60] # Tiempos que tardan los protocolos en el experimento en PsychoPy en orden
    # Se calcula el tiempo real en función del protocolo 
    protocolo= "_".join(ruta_extracción.split("/")[-1].split("_")[1:])
    for i in range(len(Tiempos_reales)):
        if protocolo == erps[i]: Tiempo_real=Tiempos_reales[i]
    Latencia= 0.02 + 0.002 + 0.016 + 0.005 # Bluethooth + leer y grabar datos + refesco pantalla + latencia SO (s)
    DeltaT=Tiempo_real-Tiempo_grabado
    #Calcular donde empieza y acaba la época, extraerla y preprocesarla
    for i in range(len(posiciones)):
        if protocolo[:2] == "p3":
            n=math.ceil(((posiciones[i]-1)-DeltaT+Latencia)*fs)-26
        if protocolo == "n400_1" or protocolo == "n400_3": 
            n=math.ceil(((posiciones[i]-1)*1.3-DeltaT+Latencia)*fs)-26
        if protocolo == "n400_2": 
            n=math.ceil(((posiciones[i]-1)*3.5+0.5-DeltaT+Latencia)*fs)-26  
        if protocolo == "n400_4":
            n=math.ceil(((posiciones[i]-1)*2+1-DeltaT+Latencia)*fs)-26
        if protocolo == "n400_5": 
            n=math.ceil(((posiciones[i]-1)*2-DeltaT+Latencia)*fs)-26
        m=n+256
        print("n es "+str(n)+" y m es "+str(m))
        if n < 0 or m > len(df): 
            print("los índices están fuera del rango permitido")
            continue
        # Extraer la época
        epoca = df.iloc[n:m+1].copy()
        # Filtro paso banda 6hz-50hz
        nyquist = 0.5 * fs
        low = 8.0 / nyquist
        high = 50.0 / nyquist
        b, a = butter(6, [low, high], btype='band')
        # Aplicar el filtro paso banda y eliminar valores por encima de 120 microvoltios
        for column in df.columns[1:]:
            epoca.loc[:,column] = filtfilt(b, a, epoca[column])
            epoca.loc[:,column] = epoca[column].apply(lambda x: x if abs(x) <= 120 else None)
        # Rechazo del canal auxiliar (canal 5)
        epoca_procesada = epoca.iloc[:, :-1]
        # Guardar el nuevo DataFrame en un archivo CSV, incluyendo el encabezado
        epoca_procesada.to_csv(ruta_extracción+"_ep"+str(i+1), index=False, header=True)
        #visualizar(ruta_extracción+"_ep"+str(i+1))

#Creo arrays de eprs y rutas
erps = ["p300_1_1","p300_2_1","p300_1_2","p300_2_2","n400_1","n400_2","p300_1_3","p300_2_3","p300_1_4","p300_2_4","n400_3","n400_4","n400_5"]
with open("C:/Users/pablo/Desktop/TFG/Experimento/Datos_participantes/Rutas excels.txt", "r",encoding="utf-8") as excels:
    rutas = excels.readlines()
rutas_excel = []
for ruta in rutas:
    rutas_excel.append(ruta.strip())


ruta_archivo='path'
erp= "_".join(ruta_archivo.split("/")[-1].split("_")[1:])
posiciones=posiciones(erp,rutas_excel[22])
extraer_epoca(ruta_archivo,posiciones,erps)